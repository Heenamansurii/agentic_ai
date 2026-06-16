from openpyxl import Workbook

def export_testcases_to_excel(data):
   workbook = Workbook()

   sheet = workbook.active

   sheet["A1"] = "Test Case ID"
   sheet["B1"] = "Scenario"
   sheet["C1"] = "Expected Result"
   row = 2
       
       
   for testcase in data:

                    sheet.cell(
                              row=row, 
                              column=1,
                              value=testcase["testcase_id"]
                             )

                    sheet.cell(
                               row=row,
                               column=2,
                               value=testcase["scenario"]
                               )

                    sheet.cell(
                                 row=row,
                                 column=3,
                                 value=testcase["expected_result"]
                                )

                    row += 1

                    workbook.save("outputs/testcases.xlsx")

